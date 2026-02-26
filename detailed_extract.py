import openpyxl

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)

GREEN = 'FFD9EAD3'
BLUE = 'FFC9DAF8'

all_matches = []

# =================== 2022 ===================
ws = wb['2022 Wheels Up']
print("=== 2022 Wheels Up ===")
# Left(green/D) = Team Bearman, Right(blue/E) = Team Larson
# Result cell color determines winner
matches_2022 = []
for row in ws.iter_rows(min_row=24, max_row=41):
    rn = row[0].row
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    if not f or not d:
        continue
    # skip headers
    if any(x in d.lower() for x in ['four ball', 'singles', 'best ball', '#']):
        continue
    f_fg = row[5].fill.fgColor.rgb if row[5].fill and row[5].fill.fgColor else '00'
    if f_fg == GREEN:
        winner = 'a'  # left/D team won
    elif f_fg == BLUE:
        winner = 'b'  # right/E team won
    elif f.upper() == 'AS':
        winner = 'as'
    else:
        winner = '??'
    
    # Determine type and course
    if rn <= 27:
        mtype, course = 'BB', '#7'
    elif rn <= 32:
        mtype, course = 'BB', '#4'
    else:
        mtype, course = 'Singles', '#2'
    
    m = {'year': 2022, 'event': 'Wheels Up', 'type': mtype, 'course': course,
         'team_a': d, 'team_b': e, 'result': f, 'winner': winner}
    matches_2022.append(m)
    print(f"  Row {rn}: {d:40s} vs {e:30s} -> {f:8s} winner={winner}")

# =================== 2023 ===================
ws = wb['2023 Black Magic']
print("\n=== 2023 Black Magic ===")
# Different layout: result in col D, team in col F (or similar)
# Let me check the actual structure
for row in ws.iter_rows(min_row=23, max_row=45):
    rn = row[0].row
    vals = []
    for i in range(8):
        if i < len(row):
            v = str(row[i].value).strip() if row[i].value else ''
            fg = row[i].fill.fgColor.rgb if row[i].fill and row[i].fill.fgColor else '00'
            vals.append((v, fg))
        else:
            vals.append(('', '00'))
    # D=col3, E=col4, F=col5, G=col6
    d, d_fg = vals[3]
    e, e_fg = vals[4]
    f, f_fg = vals[5]
    g, g_fg = vals[6] if len(vals) > 6 else ('', '00')
    if d and e:
        print(f"  Row {rn}: D={d:20s}[{d_fg}] E={e:6s}[{e_fg}] F={f:30s}[{f_fg}] G={g}[{g_fg}]")

# =================== 2024 ===================
ws = wb['2024 Wet']
print("\n=== 2024 Wet ===")
for row in ws.iter_rows(min_row=23, max_row=45):
    rn = row[0].row
    vals = []
    for i in range(8):
        if i < len(row):
            v = str(row[i].value).strip() if row[i].value else ''
            fg = row[i].fill.fgColor.rgb if row[i].fill and row[i].fill.fgColor else '00'
            vals.append((v, fg))
        else:
            vals.append(('', '00'))
    d, d_fg = vals[3]
    e, e_fg = vals[4]
    f, f_fg = vals[5]
    g, g_fg = vals[6] if len(vals) > 6 else ('', '00')
    if d:
        print(f"  Row {rn}: D={d:30s}[{d_fg}] E={e:6s}[{e_fg}] F={f:30s}[{f_fg}] G={g}[{g_fg}]")
