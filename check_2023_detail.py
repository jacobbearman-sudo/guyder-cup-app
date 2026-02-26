import openpyxl

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)
ws = wb['2023 Black Magic']

# Row 32: Clinton/Doug vs Fish/Meyer, 5&4
# B=Clinton/Doug, C=point, D=result=5&4, E=point, F=Fish/Meyer
print("=== 2023 Row 32 (Clinton/Doug vs Fish/Meyer) ===")
for row in ws.iter_rows(min_row=32, max_row=32):
    b = str(row[1].value).strip() if row[1].value else ''
    c = str(row[2].value).strip() if row[2].value else ''
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    print(f"B={b}, C={c}, D={d}, E={e}, F={f}")
    # C=left points (Team Doug), E=right points (Team AB)
    print(f"Team Doug (left) points={c}, Team AB (right) points={e}")
    if c and float(c) == 1.0:
        print("Team Doug (Clinton/Doug) WON")
    elif e and float(e) == 1.0:
        print("Team AB (Fish/Meyer) WON")

# Also check overall 2023 points
print("\n=== 2023 Overall ===")
for row in ws.iter_rows(min_row=43, max_row=43):
    vals = [str(c.value).strip() if c.value else '' for c in row[:7]]
    print(f"Row 43: {vals}")

# Re-check all 2023 BB Trevino matches
print("\n=== 2023 BB Trevino (rows 30-33) ===")
for row in ws.iter_rows(min_row=30, max_row=33):
    rn = row[0].row
    b = str(row[1].value).strip() if row[1].value else ''
    c = str(row[2].value).strip() if row[2].value else ''
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    print(f"Row {rn}: left={b} pts={c}, result={d}, pts={e}, right={f}")
    if 'Team' in b:
        continue
    if d.upper() == 'AS':
        print("  -> HALVED")
    elif c and float(c) == 1.0:
        print(f"  -> LEFT WON ({b})")
    elif e and float(e) == 1.0:
        print(f"  -> RIGHT WON ({f})")
    elif not c and not e:
        print(f"  -> NO POINTS RECORDED - need to determine winner")
