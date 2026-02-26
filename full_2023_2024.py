import openpyxl

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)

# =================== 2023 ===================
# Layout: columns B/C have left team info, D=result, E=points, F=right team name
# Need to see columns A through H
ws = wb['2023 Black Magic']
print("=== 2023 Black Magic - Full columns A-H ===")
for row in ws.iter_rows(min_row=23, max_row=43):
    rn = row[0].row
    vals = []
    for i in range(9):
        if i < len(row):
            v = str(row[i].value).strip() if row[i].value else ''
            fg = row[i].fill.fgColor.rgb if row[i].fill and row[i].fill.fgColor else '00'
        else:
            v, fg = '', '00'
        vals.append(f'{v}')
    print(f"Row {rn}: A={vals[0]:15s} B={vals[1]:15s} C={vals[2]:20s} D={vals[3]:10s} E={vals[4]:6s} F={vals[5]:30s} G={vals[6]:15s} H={vals[7]:15s}")

# =================== 2024 ===================
ws = wb['2024 Wet']
print("\n=== 2024 Wet - Full columns A-H ===")
for row in ws.iter_rows(min_row=23, max_row=44):
    rn = row[0].row
    vals = []
    for i in range(9):
        if i < len(row):
            v = str(row[i].value).strip() if row[i].value else ''
        else:
            v = ''
        vals.append(v)
    print(f"Row {rn}: A={vals[0]:20s} B={vals[1]:20s} C={vals[2]:20s} D={vals[3]:10s} E={vals[4]:6s} F={vals[5]:30s} G={vals[6]:15s} H={vals[7]:15s}")
