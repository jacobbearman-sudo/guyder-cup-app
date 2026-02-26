import openpyxl
import re

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)

GREEN = 'FFD9EAD3'
BLUE = 'FFC9DAF8'

def clean_name(s):
    s = re.sub(r'\s*\(\d+\)\s*', '', s)
    s = re.sub(r'^\s*\d+\s*', '', s)
    s = s.strip()
    s = re.sub(r'\s+\d+$', '', s)
    name_map = {
        'Jake': 'Bearman', 'Rich': 'Rames', 'Kiwi': 'AJ',
        'Jackson': 'AJ', 'Dan N.': 'Niemeyer',
        'Ab': 'AB', 'Ben': 'Larson', 'J': 'J4',
        'A B': 'AB', 'clinton': 'Clinton', 'rams': 'Rams',
        'booth': 'Booth', 'fish': 'Fish', 'Bird': 'Littel',
        'Hogan': 'Bobby',
    }
    return name_map.get(s, s)

def clean_team(s):
    parts = re.split(r'/', s)
    return '/'.join(clean_name(p.strip()) for p in parts if p.strip())

all_matches = []

# =================== 2022 Wheels Up ===================
ws = wb['2022 Wheels Up']
sections_2022 = [
    (24, 27, 'BB', '#7'),
    (29, 32, 'BB', '#4'),
    (34, 41, 'Singles', '#2'),
]
for start, end, mtype, course in sections_2022:
    for row in ws.iter_rows(min_row=start, max_row=end):
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        if not f or not d:
            continue
        f_fg = row[5].fill.fgColor.rgb if row[5].fill and row[5].fill.fgColor else '00'
        if f_fg == GREEN:
            winner = 'a'
        elif f_fg == BLUE:
            winner = 'b'
        elif f.upper() == 'AS':
            winner = 'as'
        else:
            continue
        all_matches.append({
            'year': 2022, 'event': 'Wheels Up', 'type': mtype, 'course': course,
            'team_a': clean_team(d), 'team_b': clean_team(e), 'result': f, 'winner': winner
        })

# =================== 2023 Black Magic ===================
ws = wb['2023 Black Magic']
sections_2023 = [
    (25, 28, 'BB', 'Player'),
    (30, 33, 'BB', 'Trevino'),
    (35, 42, 'Singles', 'Erin Hills'),
]
for start, end, mtype, course in sections_2023:
    for row in ws.iter_rows(min_row=start, max_row=end):
        b = str(row[1].value).strip() if row[1].value else ''
        c = str(row[2].value).strip() if row[2].value else ''
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        if not d or 'Team' in b:
            continue
        if d.upper() == 'AS':
            winner = 'as'
        elif c and float(c) == 1.0:
            winner = 'a'
        elif e and float(e) == 1.0:
            winner = 'b'
        else:
            winner = '??'
        all_matches.append({
            'year': 2023, 'event': 'Black Magic', 'type': mtype, 'course': course,
            'team_a': clean_team(b), 'team_b': clean_team(f), 'result': d, 'winner': winner
        })

# =================== 2024 Wet ===================
ws = wb['2024 Wet']
sections_2024 = [
    (25, 28, 'BB', 'Loop Red'),
    (30, 33, 'BB', 'FD'),
    (35, 42, 'Singles', 'FD'),
]
for start, end, mtype, course in sections_2024:
    for row in ws.iter_rows(min_row=start, max_row=end):
        b = str(row[1].value).strip() if row[1].value else ''
        c = str(row[2].value).strip() if row[2].value else ''
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        if not d or 'Team' in b:
            continue
        if d.upper() == 'AS':
            winner = 'as'
        elif c and float(c) == 1.0:
            winner = 'a'
        elif e and float(e) == 1.0:
            winner = 'b'
        else:
            winner = '??'
        all_matches.append({
            'year': 2024, 'event': 'Wet', 'type': mtype, 'course': course,
            'team_a': clean_team(b), 'team_b': clean_team(f), 'result': d, 'winner': winner
        })

# =================== 2025 Battle for Sarah ===================
ws = wb['2025 Battle for Sarah']
sections_2025 = [
    (24, 27, 'BB', 'Sand'),
    (29, 32, 'BB', 'Mammoth'),
    (34, 41, 'Singles', 'Sedge'),
]
for start, end, mtype, course in sections_2025:
    for row in ws.iter_rows(min_row=start, max_row=end):
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        g = str(row[6].value).strip() if row[6].value else ''
        h = str(row[7].value).strip() if row[7].value else ''
        if not f or not d:
            continue
        if f.upper() == 'AS':
            winner = 'as'
        elif e and float(e) == 1.0:
            winner = 'a'
        elif g and float(g) == 1.0:
            winner = 'b'
        else:
            winner = '??'
        all_matches.append({
            'year': 2025, 'event': 'Battle for Sarah', 'type': mtype, 'course': course,
            'team_a': clean_team(d), 'team_b': clean_team(h), 'result': f, 'winner': winner
        })

# Produce the Python literal for MATCH_HISTORY
print("MATCH_HISTORY = [")
for m in all_matches:
    print(f"    {{'year': {m['year']}, 'event': '{m['event']}', 'type': '{m['type']}', 'course': '{m['course']}', 'team_a': '{m['team_a']}', 'team_b': '{m['team_b']}', 'result': '{m['result']}', 'winner': '{m['winner']}'}},")
print("]")

# Validate record book entries
print("\n=== Individual Match Records (since 2022) ===")
players = set()
for m in all_matches:
    for p in m['team_a'].split('/') + m['team_b'].split('/'):
        players.add(p.strip())

records = {}
for p in sorted(players):
    w, l, h = 0, 0, 0
    for m in all_matches:
        a_list = [x.strip() for x in m['team_a'].split('/')]
        b_list = [x.strip() for x in m['team_b'].split('/')]
        if p in a_list:
            side = 'a'
        elif p in b_list:
            side = 'b'
        else:
            continue
        if m['winner'] == 'as':
            h += 1
        elif m['winner'] == side:
            w += 1
        else:
            l += 1
    records[p] = (w, l, h)
    total = w + 0.5 * h
    print(f"{p:15s}: {w}W-{l}L-{h}AS (Total pts: {total})")

# Record Book says:
# Fish: 6W-7L-0AS -> Total 6
# Bearman: 6W-6L-1AS -> Total 6.5
# Clinton: 9W-4L-0AS -> Total 9
print("\n=== Key comparisons ===")
for name, expected in [('Fish', '6-7-0'), ('Bearman', '6-6-1'), ('Clinton', '9-4-0'), ('Larson', '2-10-1'), ('Bobby', '7-1-2')]:
    if name in records:
        w, l, h = records[name]
        print(f"{name}: got {w}-{l}-{h}, expected {expected}")
