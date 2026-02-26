import openpyxl
import json

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)

GREEN = 'FFD9EAD3'
BLUE = 'FFC9DAF8'

all_matches = []

# =================== 2022 Wheels Up ===================
ws = wb['2022 Wheels Up']
# D=Team Bearman (green), E=Team Larson (blue)
# Result cell color: green=left won, blue=right won
sections_2022 = [
    (24, 27, 'BB', '#7'),
    (29, 32, 'BB', '#4'),
    (34, 41, 'Singles', '#2'),
]

for start, end, mtype, course in sections_2022:
    for row in ws.iter_rows(min_row=start, max_row=end):
        rn = row[0].row
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        if not f or not d:
            continue

        f_fg = row[5].fill.fgColor.rgb if row[5].fill and row[5].fill.fgColor else '00'
        if f_fg == GREEN:
            winner = 'a'
        elif f_fg == BLUE:
            winner = 'b'
        elif f.upper() == 'AS':
            winner = 'as'
        else:
            winner = '??'

        # Clean team names (remove handicap numbers in parens)
        import re
        def clean_name(s):
            s = re.sub(r'\s*\(\d+\)\s*', '', s)
            s = re.sub(r'\s*\d+\s*$', '', s)
            s = s.strip()
            name_map = {
                'Jake': 'Bearman', 'Rich': 'Rames', 'Kiwi': 'AJ',
                'Jackson': 'AJ', 'Dan N.': 'Niemeyer',
            }
            return name_map.get(s, s)

        def clean_team(s):
            parts = re.split(r'/', s)
            return '/'.join(clean_name(p.strip()) for p in parts)

        ta = clean_team(d)
        tb = clean_team(e)

        all_matches.append({
            'year': 2022, 'event': 'Wheels Up', 'type': mtype, 'course': course,
            'team_a': ta, 'team_b': tb, 'result': f, 'winner': winner
        })

# =================== 2023 Black Magic ===================
ws = wb['2023 Black Magic']
# B=Team Doug (left), F=Team AB (right)
# C=left points (1.0/0.5), E=right points
sections_2023 = [
    (25, 28, 'BB', 'Player'),
    (30, 33, 'BB', 'Trevino'),
    (35, 42, 'Singles', 'Erin Hills'),
]

import re
def clean_name_23(s):
    s = re.sub(r'\s*\(\d+\)\s*', '', s)
    s = s.strip()
    name_map = {
        'Rich': 'Rames', 'Ben': 'Larson', 'Kiwi': 'AJ',
    }
    return name_map.get(s, s)

def clean_team_23(s):
    parts = re.split(r'/', s)
    return '/'.join(clean_name_23(p.strip()) for p in parts)

for start, end, mtype, course in sections_2023:
    for row in ws.iter_rows(min_row=start, max_row=end):
        rn = row[0].row
        b = str(row[1].value).strip() if row[1].value else ''
        c = str(row[2].value).strip() if row[2].value else ''
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        if not d or 'Team' in b:
            continue

        # B = left team, F = right team
        # C = left points, E = right points
        # d = result
        ta = clean_team_23(b)
        tb = clean_team_23(f)

        if d.upper() == 'AS':
            winner = 'as'
        elif c and float(c) == 1.0:
            winner = 'a'  # left team won
        elif e and float(e) == 1.0:
            winner = 'b'  # right team won
        else:
            winner = '??'

        all_matches.append({
            'year': 2023, 'event': 'Black Magic', 'type': mtype, 'course': course,
            'team_a': ta, 'team_b': tb, 'result': d, 'winner': winner
        })

# =================== 2024 Wet ===================
ws = wb['2024 Wet']
# B=Team Dan (left), F=Team Jon (right)
# C=left points, E=right points
sections_2024 = [
    (25, 28, 'BB', 'Loop Red'),
    (30, 33, 'BB', 'FD'),
    (35, 42, 'Singles', 'FD'),
]

def clean_name_24(s):
    s = re.sub(r'\s*\(\d+\)\s*', '', s)
    s = s.strip()
    name_map = {
        'A B': 'AB', 'clinton': 'Clinton', 'rams': 'Rams',
    }
    return name_map.get(s, s)

def clean_team_24(s):
    parts = re.split(r'/', s)
    return '/'.join(clean_name_24(p.strip()) for p in parts)

for start, end, mtype, course in sections_2024:
    for row in ws.iter_rows(min_row=start, max_row=end):
        rn = row[0].row
        b = str(row[1].value).strip() if row[1].value else ''
        c = str(row[2].value).strip() if row[2].value else ''
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        if not d or 'Team' in b:
            continue

        ta = clean_team_24(b)
        tb = clean_team_24(f)

        if d.upper() == 'AS':
            winner = 'as'
        elif c and float(c) == 1.0:
            winner = 'a'
        elif e and float(e) == 1.0:
            winner = 'b'
        else:
            winner = '??'

        all_matches.append({
            'year': 2024, 'event': 'Wet', 'type': mtype, 'course': course,
            'team_a': ta, 'team_b': tb, 'result': d, 'winner': winner
        })

# =================== 2025 Battle for Sarah ===================
ws = wb['2025 Battle for Sarah']
# D=Team Rams (green, left), H=Team Rio (blue, right)
# E=left points, G=right points, F=result
sections_2025 = [
    (24, 27, 'BB', 'Sand'),
    (29, 32, 'BB', 'Mammoth'),
    (34, 41, 'Singles', 'Sedge'),
]

def clean_name_25(s):
    s = re.sub(r'\s*\(\d+\)\s*', '', s)
    s = s.strip()
    name_map = {
        'booth': 'Booth', 'fish': 'Fish', 'clinton': 'Clinton',
        'Bird': 'Littel',  # might be someone else
    }
    return name_map.get(s, s)

def clean_team_25(s):
    parts = re.split(r'/', s)
    return '/'.join(clean_name_25(p.strip()) for p in parts)

for start, end, mtype, course in sections_2025:
    for row in ws.iter_rows(min_row=start, max_row=end):
        rn = row[0].row
        d = str(row[3].value).strip() if row[3].value else ''
        e = str(row[4].value).strip() if row[4].value else ''
        f = str(row[5].value).strip() if row[5].value else ''
        g = str(row[6].value).strip() if row[6].value else ''
        h = str(row[7].value).strip() if row[7].value else ''
        if not f or not d:
            continue

        ta = clean_team_25(d)
        tb = clean_team_25(h)

        if f.upper() == 'AS':
            winner = 'as'
        elif e and float(e) == 1.0:
            winner = 'a'
        elif g and float(g) == 1.0:
            winner = 'b'
        else:
            winner = '??'

        all_matches.append({
            'year': 2025, 'event': 'Battle for Sarah', 'type': mtype, 'course': course,
            'team_a': ta, 'team_b': tb, 'result': f, 'winner': winner
        })

# Print all matches
print("MATCH_HISTORY = [")
for m in all_matches:
    print(f"    {{'year': {m['year']}, 'event': '{m['event']}', 'type': '{m['type']}', 'course': '{m['course']}', 'team_a': '{m['team_a']}', 'team_b': '{m['team_b']}', 'result': '{m['result']}', 'winner': '{m['winner']}'}},")
print("]")

# Verify Fish vs Clinton
print("\n=== Fish vs Clinton H2H ===")
wins, losses, halves = 0, 0, 0
for m in all_matches:
    a_players = [p.strip() for p in m['team_a'].split('/')]
    b_players = [p.strip() for p in m['team_b'].split('/')]
    fish_in_a = 'Fish' in a_players
    fish_in_b = 'Fish' in b_players
    clinton_in_a = 'Clinton' in a_players
    clinton_in_b = 'Clinton' in b_players
    if not (fish_in_a or fish_in_b):
        continue
    if not (clinton_in_a or clinton_in_b):
        continue
    fish_side = 'a' if fish_in_a else 'b'
    clinton_side = 'a' if clinton_in_a else 'b'
    if fish_side == clinton_side:
        print(f"  {m['year']} {m['type']}: TEAMMATES ({m['team_a']} / {m['team_b']})")
        continue
    if m['winner'] == 'as':
        halves += 1
        result = 'HALVED'
    elif m['winner'] == fish_side:
        wins += 1
        result = 'FISH WIN'
    else:
        losses += 1
        result = 'CLINTON WIN'
    print(f"  {m['year']} {m['type']}: {m['team_a']} vs {m['team_b']} -> {m['result']} winner={m['winner']} -> {result}")
print(f"\nFish vs Clinton: {wins}W-{losses}L-{halves}AS")

# Also verify Fish overall record
print("\n=== Fish Overall Match Record ===")
fw, fl, fh = 0, 0, 0
for m in all_matches:
    a_players = [p.strip() for p in m['team_a'].split('/')]
    b_players = [p.strip() for p in m['team_b'].split('/')]
    if 'Fish' in a_players:
        side = 'a'
    elif 'Fish' in b_players:
        side = 'b'
    else:
        continue
    if m['winner'] == 'as':
        fh += 1
    elif m['winner'] == side:
        fw += 1
    else:
        fl += 1
print(f"Fish: {fw}W-{fl}L-{fh}AS (Record Book says 6W-7L-0AS)")

# Verify 2022 total (should be 8-8)
print("\n=== 2022 Team Score ===")
left, right = 0, 0
for m in all_matches:
    if m['year'] != 2022:
        continue
    if m['winner'] == 'a':
        left += 1
    elif m['winner'] == 'b':
        right += 1
    else:
        left += 0.5
        right += 0.5
print(f"Team Bearman: {left}, Team Larson: {right} (should be 8-8)")
