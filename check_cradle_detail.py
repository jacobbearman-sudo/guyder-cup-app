import openpyxl
wb = openpyxl.load_workbook("/Users/jb/Desktop/The Guyder Cup.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "cradle" in cell.value.lower():
                r = cell.row
                print(f"  Found 'cradle' at {cell.coordinate}: {cell.value}")
                print(f"  Dumping rows {r-2} to {r+20}:")
                for dump_row in range(max(1, r-2), min(ws.max_row+1, r+20)):
                    vals = []
                    for c in range(1, 15):
                        v = ws.cell(row=dump_row, column=c).value
                        fill = ws.cell(row=dump_row, column=c).fill
                        fg = fill.fgColor.rgb if fill.fgColor and fill.fgColor.rgb else "none"
                        if v is not None:
                            vals.append(f"{openpyxl.utils.get_column_letter(c)}={v} (fill={fg})")
                    if vals:
                        print(f"    Row {dump_row}: {', '.join(vals)}")
                print()
