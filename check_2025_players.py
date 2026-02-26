import openpyxl

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)
ws = wb['2025 Battle for Sarah']

# Check the player list to see who's on which team
print("=== 2025 Players ===")
for row in ws.iter_rows(min_row=1, max_row=19):
    rn = row[0].row
    a = str(row[0].value).strip() if row[0].value else ''
    b = str(row[1].value).strip() if row[1].value else ''
    if a:
        fg = row[0].fill.fgColor.rgb if row[0].fill and row[0].fill.fgColor else '00'
        print(f"Row {rn}: A={a} (fill={fg}), B={b}")

# Check Bird and Hogan matches
print("\n=== 2025 Singles details ===")
for row in ws.iter_rows(min_row=34, max_row=41):
    rn = row[0].row
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    g = str(row[6].value).strip() if row[6].value else ''
    h = str(row[7].value).strip() if row[7].value else ''
    print(f"Row {rn}: {d:15s} pts={e:4s} result={f:8s} pts={g:4s} {h}")
