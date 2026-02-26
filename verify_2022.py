import openpyxl

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)

# Double-check 2022 Row 35: Clinton (3) vs Fish, 4 & 3
# Result color was FFC9DAF8 (blue = Team Larson won)
# But who is on Team Larson in 2022?
# Team Bearman (green/left): Bearman, Rames, Niemeyer, Meyer, Rams, Doug, Rio, Clinton, Rich(=Rames?)
# Team Larson (blue/right): Larson, J4, Fish, Booth, Bobby, DJ, AB, AJ

ws = wb['2022 Wheels Up']
# Let's look at row 35 color carefully
for row in ws.iter_rows(min_row=35, max_row=35):
    d = row[3]  # Clinton (3)
    e = row[4]  # Fish
    f = row[5]  # 4 & 3
    print(f"D value={d.value}, fill_fg={d.fill.fgColor.rgb if d.fill and d.fill.fgColor else 'none'}")
    print(f"E value={e.value}, fill_fg={e.fill.fgColor.rgb if e.fill and e.fill.fgColor else 'none'}")
    print(f"F value={f.value}, fill_fg={f.fill.fgColor.rgb if f.fill and f.fill.fgColor else 'none'}")

# Clinton is on green (Team Bearman), Fish is on blue (Team Larson)
# Result color FFC9DAF8 = blue = Team Larson (Fish's team) won
# So Fish beat Clinton 4&3

# But user says 2-2... Let me check: maybe there's a 5th match I'm missing?
# Or maybe the user is mistaken about the exact H2H.
# The user specifically said "he beat clinton in BB in 2022 with booth" as the correction.
# That would change 1-3 -> 2-2 if only that one match flips.
# But our correction shows 2022 BB as Fish WIN and also flips 2022 singles to Fish WIN.
# That would give 3-1 not 2-2.

# Maybe the 2022 singles wasn't actually a Fish win. Let me double-check by counting
# all the 2022 singles results and seeing if Team Bearman vs Team Larson adds up correctly.

print("\n=== 2022 Singles verification ===")
left_pts = 0
right_pts = 0
for row in ws.iter_rows(min_row=34, max_row=41):
    rn = row[0].row
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    f_fg = row[5].fill.fgColor.rgb if row[5].fill and row[5].fill.fgColor else '00'
    if not f:
        continue
    if f_fg == 'FFD9EAD3':
        w = 'LEFT'
        left_pts += 1
    elif f_fg == 'FFC9DAF8':
        w = 'RIGHT'
        right_pts += 1
    elif f.upper() == 'AS':
        w = 'HALVED'
        left_pts += 0.5
        right_pts += 0.5
    else:
        w = f'?? ({f_fg})'
    print(f"Row {rn}: {d:20s} vs {e:15s} -> {f:8s} color={f_fg} -> {w}")
print(f"Singles: Left={left_pts}, Right={right_pts}")

print("\n=== 2022 BB #7 verification ===")
left_pts = 0
right_pts = 0
for row in ws.iter_rows(min_row=24, max_row=27):
    rn = row[0].row
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    f_fg = row[5].fill.fgColor.rgb if row[5].fill and row[5].fill.fgColor else '00'
    if not f:
        continue
    if f_fg == 'FFD9EAD3':
        w = 'LEFT'
        left_pts += 1
    elif f_fg == 'FFC9DAF8':
        w = 'RIGHT'
        right_pts += 1
    else:
        w = f'?? ({f_fg})'
    print(f"Row {rn}: {d:35s} vs {e:25s} -> {f:8s} color={f_fg} -> {w}")
print(f"BB #7: Left={left_pts}, Right={right_pts}")

print("\n=== 2022 BB #4 verification ===")
left_pts = 0
right_pts = 0
for row in ws.iter_rows(min_row=29, max_row=32):
    rn = row[0].row
    d = str(row[3].value).strip() if row[3].value else ''
    e = str(row[4].value).strip() if row[4].value else ''
    f = str(row[5].value).strip() if row[5].value else ''
    f_fg = row[5].fill.fgColor.rgb if row[5].fill and row[5].fill.fgColor else '00'
    if not f:
        continue
    if f_fg == 'FFD9EAD3':
        w = 'LEFT'
        left_pts += 1
    elif f_fg == 'FFC9DAF8':
        w = 'RIGHT'
        right_pts += 1
    else:
        w = f'?? ({f_fg})'
    print(f"Row {rn}: {d:35s} vs {e:25s} -> {f:8s} color={f_fg} -> {w}")
print(f"BB #4: Left={left_pts}, Right={right_pts}")
