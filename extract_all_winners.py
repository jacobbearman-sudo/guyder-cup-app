import openpyxl

wb = openpyxl.load_workbook('/Users/jb/Desktop/The Guyder Cup.xlsx', data_only=True)

GREEN = 'FFD9EAD3'
BLUE = 'FFC9DAF8'
GREEN2 = 'FF93C47D'
BLUE2 = 'FF6D9EEB'

sheets_config = {
    '2022 Wheels Up': {'start': 24, 'end': 41, 'left_team': 'Bearman', 'right_team': 'Larson'},
    '2023 Black Magic': {'start': 1, 'end': 60, 'left_team': None, 'right_team': None},
    '2024 Wet': {'start': 1, 'end': 60, 'left_team': None, 'right_team': None},
    '2025 Battle for Sarah': {'start': 1, 'end': 60, 'left_team': None, 'right_team': None},
}

for sheet_name in wb.sheetnames:
    if sheet_name.startswith('20') and int(sheet_name[:4]) >= 2022:
        ws = wb[sheet_name]
        print(f'\n=== {sheet_name} ===')
        for row in ws.iter_rows(min_row=1, max_row=60):
            row_num = row[0].row
            if len(row) < 6:
                continue
            d_val = str(row[3].value).strip() if row[3].value else ''
            e_val = str(row[4].value).strip() if row[4].value else ''
            f_val = str(row[5].value).strip() if row[5].value else ''
            if not f_val or not d_val:
                continue
            if f_val in ['', 'None']:
                continue
            # Skip headers
            if any(x in d_val.lower() for x in ['four ball', 'singles', 'best ball', 'paid', '#']):
                if 'ball' in d_val.lower() or 'singles' in d_val.lower():
                    continue
            f_fill = row[5].fill
            f_fg = f_fill.fgColor.rgb if f_fill and f_fill.fgColor else '00000000'
            d_fill = row[3].fill
            d_fg = d_fill.fgColor.rgb if d_fill and d_fill.fgColor else '00000000'
            
            if f_fg == GREEN:
                winner_side = 'left'
            elif f_fg == BLUE:
                winner_side = 'right'
            elif f_val.upper() == 'AS':
                winner_side = 'halved'
            else:
                winner_side = f'UNKNOWN(fg={f_fg})'
            
            print(f'Row {row_num}: {d_val:40s} vs {e_val:40s} -> {f_val:8s} winner={winner_side}')
