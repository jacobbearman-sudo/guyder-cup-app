import openpyxl
wb = openpyxl.load_workbook("/Users/jb/Desktop/The Guyder Cup.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "cradle" in cell.value.lower():
                print(f"Sheet: {sheet_name}, Cell: {cell.coordinate}, Value: {cell.value}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "tiebreak" in cell.value.lower():
                print(f"Sheet: {sheet_name}, Cell: {cell.coordinate}, Value: {cell.value}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "tie" in cell.value.lower():
                print(f"Sheet: {sheet_name}, Cell: {cell.coordinate}, Value: {cell.value}")
